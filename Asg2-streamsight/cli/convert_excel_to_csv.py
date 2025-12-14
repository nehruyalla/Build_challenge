#!/usr/bin/env python3
"""Convert Excel files to CSV for streaming processing.

This script converts an Excel file to CSV format, which is required for
the streaming pipeline. Uses openpyxl in read-only mode for memory efficiency.
"""

import sys
from pathlib import Path

from openpyxl import load_workbook


def convert_excel_to_csv(excel_path: Path, csv_path: Path) -> None:
    """Convert an Excel file to CSV format.

    Args:
        excel_path: Path to input Excel file
        csv_path: Path to output CSV file

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If Excel file has no data
    """
    print(f"Converting {excel_path} to {csv_path}...")

    if not excel_path.exists():
        msg = f"Excel file not found: {excel_path}"
        raise FileNotFoundError(msg)

    # Load workbook in read-only mode for memory efficiency
    wb = load_workbook(filename=excel_path, read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        msg = "Workbook has no active sheet"
        raise ValueError(msg)

    # Write to CSV
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    row_count = 0

    with open(csv_path, "w", encoding="utf-8") as f:
        for row in ws.iter_rows(values_only=True):
            # Skip empty rows
            if all(cell is None for cell in row):
                continue

            # Convert row to CSV format
            # Handle None values and escape commas in strings
            csv_row = []
            for cell in row:
                if cell is None:
                    csv_row.append("")
                elif isinstance(cell, str):
                    # Escape quotes and wrap in quotes if contains comma
                    if "," in cell or '"' in cell or "\n" in cell:
                        escaped = cell.replace('"', '""')
                        csv_row.append(f'"{escaped}"')
                    else:
                        csv_row.append(cell)
                else:
                    csv_row.append(str(cell))

            f.write(",".join(csv_row) + "\n")
            row_count += 1

    wb.close()

    print(f"Converted {row_count:,} rows")
    print(f"Output: {csv_path}")


def main() -> None:
    """Main entry point."""
    # Default paths
    project_root = Path(__file__).parent.parent
    excel_path = project_root / "dataset" / "Online Retail.xlsx"
    csv_path = project_root / "dataset" / "Online_Retail.csv"

    # Allow command-line override
    if len(sys.argv) >= 2:
        excel_path = Path(sys.argv[1])
    if len(sys.argv) >= 3:
        csv_path = Path(sys.argv[2])

    try:
        convert_excel_to_csv(excel_path, csv_path)
        print("\nConversion successful!")
    except Exception as e:
        print(f"\nError: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()

